"""
Excel Exporter Module.

This module provides Excel file generation for invoice extraction
results. Uses openpyxl for modern Excel format support.

Features:
    - Formatted headers
    - Auto-column width
    - Metadata sheet
    - Multiple result support

Author: ML Engineering Team
"""

from pathlib import Path
from typing import List, Optional, Union, Dict, Any
from datetime import datetime

from config import get_config
from src.utils.logger import get_logger
from src.utils.helpers import ensure_directory, generate_timestamp
from src.utils.exceptions import ExcelExportError
from src.model_inference.extraction_result import ExtractionResult

# Initialize module logger
logger = get_logger(__name__)


class ExcelExporter:
    """
    Exports extraction results to Excel format.
    
    Creates well-formatted Excel files with extraction results,
    including headers, data formatting, and optional metadata.
    
    Attributes:
        output_dir: Directory for output files
        include_metadata: Whether to include metadata sheet
        include_confidence: Whether to include confidence scores
        
    Example:
        >>> exporter = ExcelExporter()
        >>> filepath = exporter.export(results, "extractions.xlsx")
        >>> print(f"Saved to: {filepath}")
    """
    
    # Column definitions
    COLUMNS = [
        ('Invoice Number', 'invoice_number'),
        ('Invoice Date', 'invoice_date'),
        ('Vendor Name', 'vendor_name'),
        ('Customer Name', 'customer_name'),
        ('Total Amount', 'total_amount'),
        ('Payment Due Date', 'payment_due_date'),
    ]
    
    METADATA_COLUMNS = [
        ('Source File', 'source_file'),
        ('Extraction Timestamp', 'extraction_timestamp'),
        ('Model Name', 'model_name'),
        ('Processing Time (s)', 'processing_time'),
        ('Success', 'success'),
        ('Extraction Rate (%)', 'extraction_rate'),
        ('Average Confidence', 'average_confidence'),
    ]
    
    def __init__(self) -> None:
        """Initialize the Excel exporter with configuration."""
        self.output_dir = Path(get_config("paths.output_dir", "outputs"))
        self.include_metadata = get_config("output.excel.include_metadata", True)
        self.include_confidence = get_config("output.excel.include_confidence", True)
        self.sheet_name = get_config("output.excel.sheet_name", "Extracted Data")
        
        # Check for openpyxl
        self._check_dependencies()
        
        logger.debug(f"ExcelExporter initialized (output_dir: {self.output_dir})")
    
    def _check_dependencies(self) -> None:
        """Check if required libraries are available."""
        try:
            import openpyxl
            self._openpyxl = openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )
    
    def export(
        self,
        results: Union[ExtractionResult, List[ExtractionResult]],
        filename: Optional[str] = None,
        output_dir: Optional[str] = None
    ) -> str:
        """
        Export extraction results to Excel file.
        
        Args:
            results: Single result or list of results to export.
            filename: Output filename. If None, auto-generated.
            output_dir: Output directory. If None, uses configured dir.
            
        Returns:
            Path to the created Excel file.
            
        Raises:
            ExcelExportError: If export fails.
            
        Example:
            >>> path = exporter.export(results, "invoice_data.xlsx")
        """
        # Normalize to list
        if isinstance(results, ExtractionResult):
            results = [results]
        
        if not results:
            raise ExcelExportError("No results", "No results to export")
        
        # Determine output path
        if output_dir:
            out_dir = Path(output_dir)
        else:
            out_dir = self.output_dir
        
        ensure_directory(out_dir)
        
        if filename is None:
            timestamp = generate_timestamp()
            filename = f"invoice_extractions_{timestamp}.xlsx"
        
        filepath = out_dir / filename
        
        try:
            # Create workbook
            workbook = self._openpyxl.Workbook()
            
            # Create main data sheet
            self._create_data_sheet(workbook, results)
            
            # Create metadata sheet if enabled
            if self.include_metadata:
                self._create_metadata_sheet(workbook, results)
            
            # Create confidence sheet if enabled
            if self.include_confidence:
                self._create_confidence_sheet(workbook, results)
            
            # Save workbook
            workbook.save(filepath)
            
            logger.info(f"Excel file saved: {filepath} ({len(results)} records)")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise ExcelExportError(str(filepath), str(e))
    
    def _create_data_sheet(
        self,
        workbook,
        results: List[ExtractionResult]
    ) -> None:
        """
        Create the main data sheet with extraction results.
        
        Args:
            workbook: openpyxl Workbook instance.
            results: List of extraction results.
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get active sheet and rename
        sheet = workbook.active
        sheet.title = self.sheet_name
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, (header_name, _) in enumerate(self.COLUMNS, 1):
            cell = sheet.cell(row=1, column=col, value=header_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_num, result in enumerate(results, 2):
            for col, (_, field_name) in enumerate(self.COLUMNS, 1):
                value = getattr(result, field_name, '') or ''
                cell = sheet.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
        
        # Adjust column widths
        for col, (header_name, _) in enumerate(self.COLUMNS, 1):
            column_letter = get_column_letter(col)
            
            # Calculate max width
            max_length = len(header_name)
            for row in range(2, len(results) + 2):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with padding
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
    
    def _create_metadata_sheet(
        self,
        workbook,
        results: List[ExtractionResult]
    ) -> None:
        """
        Create a metadata sheet with processing information.
        
        Args:
            workbook: openpyxl Workbook instance.
            results: List of extraction results.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        sheet = workbook.create_sheet(title="Metadata")
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        
        # Write headers
        for col, (header_name, _) in enumerate(self.METADATA_COLUMNS, 1):
            cell = sheet.cell(row=1, column=col, value=header_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_num, result in enumerate(results, 2):
            for col, (_, field_name) in enumerate(self.METADATA_COLUMNS, 1):
                if field_name == 'extraction_rate':
                    value = f"{result.extraction_rate:.1f}"
                elif field_name == 'average_confidence':
                    value = f"{result.average_confidence:.2f}"
                elif field_name == 'processing_time':
                    value = f"{result.processing_time:.2f}"
                else:
                    value = getattr(result, field_name, '') or ''
                
                sheet.cell(row=row_num, column=col, value=value)
        
        # Adjust column widths
        for col in range(1, len(self.METADATA_COLUMNS) + 1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 20
    
    def _create_confidence_sheet(
        self,
        workbook,
        results: List[ExtractionResult]
    ) -> None:
        """
        Create a sheet with confidence scores for each field.
        
        Args:
            workbook: openpyxl Workbook instance.
            results: List of extraction results.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        sheet = workbook.create_sheet(title="Confidence Scores")
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        
        # Columns: Source File + all field confidence scores
        columns = ['Source File'] + [f"{name} Conf." for name, _ in self.COLUMNS]
        
        # Write headers
        for col, header in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_num, result in enumerate(results, 2):
            sheet.cell(row=row_num, column=1, value=result.source_file or '')
            
            for col, (_, field_name) in enumerate(self.COLUMNS, 2):
                confidence = result.confidence_scores.get(field_name, 0)
                sheet.cell(row=row_num, column=col, value=f"{confidence:.2f}")
        
        # Adjust column widths
        for col in range(1, len(columns) + 1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 18
    
    def get_default_filename(self) -> str:
        """
        Generate a default filename with timestamp.
        
        Returns:
            Default filename string.
        """
        timestamp = generate_timestamp()
        pattern = get_config(
            "output.excel.filename_pattern",
            "invoice_extractions_{timestamp}.xlsx"
        )
        return pattern.format(timestamp=timestamp)
